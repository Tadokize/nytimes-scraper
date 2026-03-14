import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def save_to_excel(results: list, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'News'

    headers = [
        'Title',
        'Date',
        'Description',
        'Image Filename',
        'Phrase Count',
        'Has Money'
    ]

    # estilo do cabeçalho
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='1F4E79')
    header_alignment = Alignment(horizontal='center', vertical='center')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # preenche as linhas com os dados coletados
    for row, item in enumerate(results, 2):
        ws.cell(row=row, column=1, value=item['title'])
        ws.cell(row=row, column=2, value=item['date'])
        ws.cell(row=row, column=3, value=item['description'])
        ws.cell(row=row, column=4, value=item['image_filename'])
        ws.cell(row=row, column=5, value=item['phrase_count'])
        ws.cell(row=row, column=6, value=item['has_money'])

    # ajusta largura das colunas para melhor leitura
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 70
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    wb.save(output_path)
    print(f'Excel salvo em: {output_path}')